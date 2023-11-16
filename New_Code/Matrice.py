## fonction qui met sous forme les dictionnaire les matrices de score

import openpyxl
from pathlib import Path
import pandas as pd

PATH = "C:/Users/Guillaume/Desktop/hERG/New_Code/"
FILE = "matrices.xlsx"
xlsx_file = Path(PATH,FILE)


def matrices(FILE):
    L = []
    wb = openpyxl.load_workbook(FILE)
    for ws in wb._sheets:
        matrice = {}
        i = 0
        for row in ws.iter_rows():
            i+=1
            if i==1:
                for cell in row[1:]:
                    matrice[cell.value.strip()]={}
            elif i==22:
                break
            else:
                j=0
                aa = ""
                for cell in row:
                    j += 1
                    if j == 1:
                        aa = cell.value.strip()
                    else:
                        matrice[ws[cell.column_letter+"1"].value.strip()][aa] = cell.value
        L.append(matrice)
    return L

def read_xlsx(xlsx_name):
    """ From a xlsx file containing 1 mutation per row ("E344G" as an example), extracts the contente and retunrs 
        a 3-column dataframe containing the AA concerned by the position, its position and the AA newly obtained   """
    input_df = pd.read_excel(xlsx_name, header=None)
    nb_mut = len(input_df.columns)
    pre_aa, pos, post_aa = [], [], []
    c1, c2, c3, c4 = [], [], [], []
    clash, mini = [], []
    total, AA_imp, nb_AA_imp = [], [], []
    for i in range(1,nb_mut):
        pre_aa.append(input_df[i][0][0])
        pos.append(input_df[i][0][1:-1]) 
        post_aa.append(input_df[i][0][-1:])
        c1.append(input_df[i][1])
        c2.append(input_df[i][2])
        c3.append(input_df[i][3])
        clash.append(input_df[i][4])
        mini.append(input_df[i][5])
        c4.append(input_df[i][6])
        total.append(input_df[i][7])
        AA_imp.append(input_df[i][8])
        nb_AA_imp.append(input_df[i][9])
    mutations_read=pd.DataFrame()
    mutations_read["pre_aa"] = pre_aa 
    mutations_read["pos"] = pos 
    mutations_read["post_aa"] = post_aa 
    return pre_aa,pos,post_aa,c1,c2,c3,clash,mini,c4,total,AA_imp,nb_AA_imp

def write_xlsx(xlsx_name,mutation,critere,value):
    wb = openpyxl.load_workbook(xlsx_name)
    ws = wb._sheets[0]
    rows = [cell.value for cell in ws[1]]
    colonne = rows.index(mutation)
    column = [cell.value for cell in ws["A"]]
    ligne = column.index(critere)
    if type(value) != int and type(value) != float:
        value = str(value)
    ws[ligne+1][colonne].value = value
    wb.save(xlsx_name)
    
    
    
    





    
