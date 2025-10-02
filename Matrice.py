## fonction qui met sous forme les dictionnaire les matrices de score

import openpyxl
from pathlib import Path
import pandas as pd

def matrices(FILE):
    L = []
    wb = openpyxl.load_workbook(FILE)
    for ws in wb._sheets:
        matrice = {}
        i = 0
        for row in ws.iter_rows():
            i+=1
            if i==1:
                for cell in row[1:]:
                    matrice[cell.value.strip()]={}
            elif i==22:
                break
            else:
                j=0
                aa = ""
                for cell in row:
                    j += 1
                    if j == 1:
                        aa = cell.value.strip()
                    else:
                        matrice[ws[cell.column_letter+"1"].value.strip()][aa] = cell.value
        L.append(matrice)
    return L

def clean_xlsx(xlsx_name):
    wb = openpyxl.load_workbook(xlsx_name)
    ws = wb._sheets[0]
    rows = [cell.value for cell in ws[1]]
    for mut in rows:
        colonne = rows.index(mut)
        if mut != None:
            ws[1][colonne].value = mut.strip()
    wb.save(xlsx_name)
        
    
def read_xlsx(xlsx_name):
    """ From a xlsx file containing 1 mutation per row ("E344G" as an example), extracts the contente and retunrs 
        a 3-column dataframe containing the AA concerned by the position, its position and the AA newly obtained   """
    input_df = pd.read_excel(xlsx_name, header=None)
    nb_mut = len(input_df.columns)
    pre_aa, pos, post_aa = [], [], []
    c1, c2, c3, c4, c5 = [], [], [], [], []
    clash, mini = [], []
    total, AA_con, nb_AA_con, AA_imp, nb_AA_imp = [], [], [], [], []
    pr1,ncr1,ncir1 = [], [], []
    pr2,ncr2,ncir2 = [], [], []
    pr3,ncr3,ncir3 = [], [], []
    pr4,ncr4,ncir4 = [], [], []
    pr5,ncr5,ncir5 = [], [], []
    final_score,sps = [], []
    for i in range(1,nb_mut):
        mut = input_df[i][0].strip()
        pre_aa.append(mut[0])
        pos.append(mut[1:-1]) 
        post_aa.append(mut[-1:])
        c1.append(input_df[i][1])
        c2.append(input_df[i][2])
        c3.append(input_df[i][3])
        clash.append(input_df[i][4])
        mini.append(input_df[i][5])
        c4.append(input_df[i][6])
        total.append(input_df[i][7])
        AA_con.append(input_df[i][8])
        nb_AA_con.append(input_df[i][9])
        AA_imp.append(input_df[i][10])
        nb_AA_imp.append(input_df[i][11])
        pr1.append(input_df[i][12])
        ncr1.append(input_df[i][13])
        ncir1.append(input_df[i][14])
        pr2.append(input_df[i][15])
        ncr2.append(input_df[i][16])
        ncir2.append(input_df[i][17])
        pr3.append(input_df[i][18])
        ncr3.append(input_df[i][19])
        ncir3.append(input_df[i][20])
        pr4.append(input_df[i][21])
        ncr4.append(input_df[i][22])
        ncir4.append(input_df[i][23])
        pr5.append(input_df[i][24])
        ncr5.append(input_df[i][25])
        ncir5.append(input_df[i][26])
        c5.append(input_df[i][27])
        final_score.append(input_df[i][28])
        sps.append(input_df[i][29])
    mutations_read=pd.DataFrame()
    mutations_read["pre_aa"] = pre_aa 
    mutations_read["pos"] = pos 
    mutations_read["post_aa"] = post_aa 
    return pre_aa,pos,post_aa,c1,c2,c3,clash,mini,c4,total,AA_con,nb_AA_con,AA_imp,nb_AA_imp,pr1,ncr1,ncir1,pr2,ncr2,ncir2,pr3,ncr3,ncir3,pr4,ncr4,ncir4,pr5,ncr5,ncir5,c5,final_score,sps

def write_xlsx(xlsx_name,mutation,critere,value):
    wb = openpyxl.load_workbook(xlsx_name)
    ws = wb._sheets[0]
    rows = [cell.value for cell in ws[1]]
    colonne = rows.index(mutation)
    column = [cell.value for cell in ws["A"]]
    ligne = column.index(critere)
    if type(value) != int and type(value) != float:
        value = str(value)
    ws[ligne+1][colonne].value = value
    wb.save(xlsx_name)
    
    
    
    





    
