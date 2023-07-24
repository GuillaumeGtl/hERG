import openpyxl
from pathlib import Path

PATH = "C:/Users/Guillaume/Desktop/hERG/Code/"
FILE = "matrice.xlsx"
xlsx_file = Path(PATH,FILE)
wb = openpyxl.load_workbook(xlsx_file)
sheet = wb.active
L = []
for ws in wb._sheets:
    matrice = {}
    i = 0
    for row in ws.iter_rows():
        i+=1
        if i==1:
            for cell in row[1:]:
                matrice[cell.value]={}
            print(matrice)
        else:
            j = 0
            aa = ""
            for cell in row:
                j += 1
                if j == 1:
                    aa = cell.value
                else :
                    matrice[ws[cell.column_letter+"1"].value][aa] = cell.value
    L.append(matrice)
table1,table2,table3 = L
